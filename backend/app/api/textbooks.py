from __future__ import annotations

import os
import io

from flask import Blueprint, current_app, jsonify, request
from sqlalchemy import delete, insert, select, update
from sqlalchemy.exc import SQLAlchemyError

from app.db import get_db, get_session
from app.services.deepseek import get_deepseek_client
from openpyxl import load_workbook

textbooks_bp = Blueprint("textbooks", __name__)


def _table(name: str):
    db = get_db(current_app)
    if name not in db.metadata.tables:
        db.metadata.reflect(bind=db.engine, only=[name])
    return db.metadata.tables[name]


def _build_chapter_tree(rows: list[dict]) -> list[dict]:
    by_id: dict[int, dict] = {}
    children_map: dict[int | None, list[dict]] = {}

    for r in rows:
        node = dict(r)
        node["children"] = []
        by_id[node["chapter_id"]] = node
        parent_id = node.get("parent_chapter_id")
        if parent_id in [0, "0"]:
            parent_id = None
        children_map.setdefault(parent_id, []).append(node)

    for parent_id, children in children_map.items():
        children.sort(key=lambda x: (x.get("chapter_sort") is None, x.get("chapter_sort") or 0, x["chapter_id"]))
        if parent_id is None:
            continue
        parent = by_id.get(parent_id)
        if parent is not None:
            parent["children"] = children

    roots = children_map.get(None, [])
    roots.sort(key=lambda x: (x.get("chapter_sort") is None, x.get("chapter_sort") or 0, x["chapter_id"]))
    return roots


def _chapter_summary_path(chapter_id: int) -> str:
    base = os.path.join(current_app.config["UPLOAD_DIR"], "chapter_summaries")
    os.makedirs(base, exist_ok=True)
    return os.path.join(base, f"{chapter_id}.txt")


@textbooks_bp.get("")
def list_textbooks():
    subject_id = request.args.get("subject_id", type=int)
    t = _table("textbook")

    stmt = select(
        t.c.textbook_id,
        t.c.subject_id,
        t.c.textbook_name,
        t.c.author,
        t.c.publisher,
        t.c.edition,
        t.c.create_time,
        t.c.update_time,
    )
    if subject_id is not None:
        stmt = stmt.where(t.c.subject_id == subject_id)
    stmt = stmt.order_by(t.c.textbook_id.desc())

    try:
        rows = get_session(current_app).execute(stmt).mappings().all()
        return jsonify({"items": [dict(r) for r in rows]})
    except SQLAlchemyError as err:
        return jsonify({"error": {"message": str(err), "type": err.__class__.__name__}}), 500


@textbooks_bp.post("")
def create_textbook():
    payload = request.get_json(silent=True) or {}
    t = _table("textbook")

    data = {
        "subject_id": payload.get("subject_id"),
        "textbook_name": payload.get("textbook_name"),
        "author": payload.get("author"),
        "publisher": payload.get("publisher"),
        "edition": payload.get("edition"),
    }

    if not data["subject_id"] or not data["textbook_name"]:
        return jsonify({"error": {"message": "subject_id 和 textbook_name 必填", "type": "BadRequest"}}), 400

    try:
        session = get_session(current_app)
        res = session.execute(insert(t).values(**data))
        session.commit()
        textbook_id = res.inserted_primary_key[0] if res.inserted_primary_key else None
        return jsonify({"textbook_id": textbook_id})
    except SQLAlchemyError as err:
        get_session(current_app).rollback()
        return jsonify({"error": {"message": str(err), "type": err.__class__.__name__}}), 500


@textbooks_bp.get("/<int:textbook_id>")
def get_textbook(textbook_id: int):
    t = _table("textbook")
    stmt = select(t).where(t.c.textbook_id == textbook_id)
    try:
        row = get_session(current_app).execute(stmt).mappings().first()
        if row is None:
            return jsonify({"error": {"message": "教材不存在", "type": "NotFound"}}), 404
        return jsonify({"item": dict(row)})
    except SQLAlchemyError as err:
        return jsonify({"error": {"message": str(err), "type": err.__class__.__name__}}), 500


@textbooks_bp.put("/<int:textbook_id>")
def update_textbook(textbook_id: int):
    payload = request.get_json(silent=True) or {}
    t = _table("textbook")

    data = {
        k: payload.get(k)
        for k in ["subject_id", "textbook_name", "author", "publisher", "edition"]
        if k in payload
    }
    if not data:
        return jsonify({"ok": True})

    try:
        session = get_session(current_app)
        session.execute(update(t).where(t.c.textbook_id == textbook_id).values(**data))
        session.commit()
        return jsonify({"ok": True})
    except SQLAlchemyError as err:
        get_session(current_app).rollback()
        return jsonify({"error": {"message": str(err), "type": err.__class__.__name__}}), 500


@textbooks_bp.delete("/<int:textbook_id>")
def delete_textbook(textbook_id: int):
    t = _table("textbook")
    try:
        session = get_session(current_app)
        session.execute(delete(t).where(t.c.textbook_id == textbook_id))
        session.commit()
        return jsonify({"ok": True})
    except SQLAlchemyError as err:
        get_session(current_app).rollback()
        return jsonify({"error": {"message": str(err), "type": err.__class__.__name__}}), 500


@textbooks_bp.get("/<int:textbook_id>/chapters")
def list_chapters(textbook_id: int):
    ch = _table("textbook_chapter")
    stmt = (
        select(
            ch.c.chapter_id,
            ch.c.textbook_id,
            ch.c.chapter_name,
            ch.c.chapter_level,
            ch.c.parent_chapter_id,
            ch.c.chapter_sort,
            ch.c.create_time,
        )
        .where(ch.c.textbook_id == textbook_id)
        .order_by(ch.c.chapter_sort.asc(), ch.c.chapter_id.asc())
    )
    try:
        rows = get_session(current_app).execute(stmt).mappings().all()
        items = [dict(r) for r in rows]
        return jsonify({"items": items, "tree": _build_chapter_tree(items)})
    except SQLAlchemyError as err:
        return jsonify({"error": {"message": str(err), "type": err.__class__.__name__}}), 500


@textbooks_bp.post("/<int:textbook_id>/chapters")
def create_chapter(textbook_id: int):
    payload = request.get_json(silent=True) or {}
    ch = _table("textbook_chapter")

    data = {
        "textbook_id": textbook_id,
        "chapter_name": payload.get("chapter_name"),
        "chapter_level": payload.get("chapter_level"),
        "parent_chapter_id": payload.get("parent_chapter_id"),
        "chapter_sort": payload.get("chapter_sort"),
    }

    if not data["chapter_name"] or data["chapter_level"] is None:
        return jsonify({"error": {"message": "chapter_name 和 chapter_level 必填", "type": "BadRequest"}}), 400

    try:
        session = get_session(current_app)
        res = session.execute(insert(ch).values(**data))
        session.commit()
        chapter_id = res.inserted_primary_key[0] if res.inserted_primary_key else None
        return jsonify({"chapter_id": chapter_id})
    except SQLAlchemyError as err:
        get_session(current_app).rollback()
        return jsonify({"error": {"message": str(err), "type": err.__class__.__name__}}), 500


@textbooks_bp.post("/<int:textbook_id>/chapters/import/excel")
def import_chapters_excel(textbook_id: int):
    file = request.files.get("file")
    if file is None:
        return jsonify({"error": {"message": "缺少上传文件 file", "type": "BadRequest"}}), 400

    wb = load_workbook(filename=io.BytesIO(file.read()), data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    header = [str(c.value).strip() if c.value is not None else "" for c in header_row]
    idx = {h: i for i, h in enumerate(header) if h}

    def pick(row, *names):
        for n in names:
            if n in idx:
                v = row[idx[n]].value
                if v is not None and str(v).strip() != "":
                    return v
        return None

    ch = _table("textbook_chapter")
    session = get_session(current_app)
    inserted = 0
    skipped = 0

    try:
        for row in ws.iter_rows(min_row=2):
            name = pick(row, "chapter_name", "章节名称", "chapter")
            if name is None or str(name).strip() == "":
                skipped += 1
                continue

            level = pick(row, "chapter_level", "层级", "level")
            sort = pick(row, "chapter_sort", "排序", "sort")
            parent_id = pick(row, "parent_chapter_id", "父章节ID", "parent_id")

            def to_int(x):
                try:
                    if x is None or str(x).strip() == "":
                        return None
                    return int(x)
                except Exception:
                    return None

            data = {
                "textbook_id": textbook_id,
                "chapter_name": str(name).strip(),
                "chapter_level": to_int(level) or 1,
                "parent_chapter_id": to_int(parent_id),
                "chapter_sort": to_int(sort) or 1,
            }

            session.execute(insert(ch).values(**data))
            inserted += 1

        session.commit()
        return jsonify({"ok": True, "inserted": inserted, "skipped": skipped})
    except SQLAlchemyError as err:
        session.rollback()
        return jsonify({"error": {"message": str(err), "type": err.__class__.__name__}}), 500


@textbooks_bp.put("/chapters/<int:chapter_id>")
def update_chapter(chapter_id: int):
    payload = request.get_json(silent=True) or {}
    ch = _table("textbook_chapter")

    data = {
        k: payload.get(k)
        for k in ["chapter_name", "chapter_level", "parent_chapter_id", "chapter_sort"]
        if k in payload
    }
    if not data:
        return jsonify({"ok": True})

    try:
        session = get_session(current_app)
        session.execute(update(ch).where(ch.c.chapter_id == chapter_id).values(**data))
        session.commit()
        return jsonify({"ok": True})
    except SQLAlchemyError as err:
        get_session(current_app).rollback()
        return jsonify({"error": {"message": str(err), "type": err.__class__.__name__}}), 500


@textbooks_bp.delete("/chapters/<int:chapter_id>")
def delete_chapter(chapter_id: int):
    ch = _table("textbook_chapter")
    try:
        session = get_session(current_app)
        session.execute(delete(ch).where(ch.c.chapter_id == chapter_id))
        session.commit()
        return jsonify({"ok": True})
    except SQLAlchemyError as err:
        get_session(current_app).rollback()
        return jsonify({"error": {"message": str(err), "type": err.__class__.__name__}}), 500


@textbooks_bp.get("/chapters/<int:chapter_id>/summary")
def get_chapter_summary(chapter_id: int):
    path = _chapter_summary_path(chapter_id)
    if not os.path.exists(path):
        return jsonify({"chapter_id": chapter_id, "summary": ""})
    with open(path, "r", encoding="utf-8") as f:
        return jsonify({"chapter_id": chapter_id, "summary": f.read()})


@textbooks_bp.put("/chapters/<int:chapter_id>/summary")
def update_chapter_summary(chapter_id: int):
    payload = request.get_json(silent=True) or {}
    summary = payload.get("summary")
    if summary is None:
        return jsonify({"error": {"message": "summary 必填", "type": "BadRequest"}}), 400
    path = _chapter_summary_path(chapter_id)
    with open(path, "w", encoding="utf-8") as f:
        f.write(str(summary))
    return jsonify({"ok": True})


@textbooks_bp.post("/chapters/<int:chapter_id>/summary/generate")
def generate_chapter_summary(chapter_id: int):
    ch = _table("textbook_chapter")
    qb = _table("question_bank")
    try:
        session = get_session(current_app)
        chapter = session.execute(select(ch).where(ch.c.chapter_id == chapter_id)).mappings().first()
        if chapter is None:
            return jsonify({"error": {"message": "章节不存在", "type": "NotFound"}}), 404

        q_stmt = (
            select(qb.c.question_content, qb.c.question_answer, qb.c.question_analysis)
            .where(qb.c.chapter_id == chapter_id)
            .where(qb.c.review_status == 1)
            .order_by(qb.c.question_id.desc())
            .limit(30)
        )
        questions = session.execute(q_stmt).mappings().all()

        system_prompt = "你是教学助理，负责为教材章节生成简洁、结构化的章节概要。"
        user_prompt = (
            "请为以下章节生成概要，要求：\n"
            "1) 200-400字\n"
            "2) 分点描述关键知识点与常见考点\n"
            "3) 不要编造不存在的章节信息\n\n"
            f"章节名称：{chapter['chapter_name']}\n"
            f"章节层级：{chapter.get('chapter_level')}\n\n"
            "参考题目（题干/答案/解析，可能不完整）：\n"
        )
        for i, q in enumerate(questions, start=1):
            user_prompt += f"\n{i}. 题干：{q.get('question_content')}\n   答案：{q.get('question_answer')}\n   解析：{q.get('question_analysis')}\n"

        client = get_deepseek_client()
        summary = client.chat(system_prompt=system_prompt, user_prompt=user_prompt, temperature=0.2)

        path = _chapter_summary_path(chapter_id)
        with open(path, "w", encoding="utf-8") as f:
            f.write(summary)
        return jsonify({"chapter_id": chapter_id, "summary": summary})
    except Exception as err:
        return jsonify({"error": {"message": str(err), "type": err.__class__.__name__}}), 500
